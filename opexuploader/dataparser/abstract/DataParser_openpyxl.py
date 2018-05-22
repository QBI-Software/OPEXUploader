# -*- coding: utf-8 -*-
"""
Utility script: DataParser - Abstract class
Reads an excel or csv file with data and extracts per subject
run from console/terminal with (example):
>python DataParser.py --filedir "data" --sheet "Sheetname_to_extract"

Created on Thu Mar 2 2017

@author: Liz Cooper-Williams, QBI
"""

import glob
from datetime import datetime
from os import R_OK, access
from os.path import join, basename, splitext, dirname, abspath
from openpyxl import load_workbook
from opexuploader.utils import findResourceDir
import pandas
import logging
from config.dbquery import DBI

class DataParser(object):
    def __init__(self, datafile=None, sheet=0,skiplines=0, header=None, etype=None):
        self.datafile = datafile #full pathname to data file
        self.resource_dir = findResourceDir()
        configdb = join(self.resource_dir, 'opexconfig.db')
        if not access(configdb,R_OK):
            print(configdb)
            raise IOError('Cannot find config database')
        try:
            self.dbi = DBI(configdb)
            if etype is not None:
                self.info = self.dbi.getInfo(etype)
                self.fields = self.dbi.getFields(etype)
            else:
                self.info = None
                self.fields = None
            self.subjects = None
            #self.incorrect = self.dbi.getIDs()
            if (self.datafile is not None and len(self.datafile)> 0):
                (bname, extn)= splitext(basename(self.datafile))
                self.ftype = extn #extension - xlsx or csv
                self.sheet = sheet
                self.skiplines = skiplines
                self.header = header
                self._loadData()
        except Exception as e:
            raise e

    def getInfoFromFile(self, etype):
        # Read expt info
        info = None
        try:
            opex = pandas.read_csv(join(self.resource_dir, 'opex.csv'))
            info = opex[opex['Expt'] == etype]
        except Exception as e:
            raise ValueError("Unable to get expt info from file", e)
        return info

    def getIdsFromFile(self):
        # Read expt info
        info = None
        try:
            info = pandas.read_csv(join(self.resource_dir, 'incorrectIds.csv'))
        except Exception as e:
            raise ValueError("Unable to get ids from file", e)
        return info

    # def __findResourcesdir(self):
    #     resource_dir = glob.glob(join(dirname(__file__), "resources"))
    #     middir = ".."
    #     ctr = 1
    #     while len(resource_dir) <= 0 and ctr < 5:
    #         resource_dir = glob.glob(join(dirname(__file__), middir, "resources"))
    #         middir = join(middir, "..")
    #         ctr += 1
    #     return abspath(resource_dir[0])

    def __checkSID(self,sid):
        """
        Replace known incorrect IDs from db
        :param sid:
        :return:
        """
        if self.dbi is not None:
            rsid = self.dbi.getCorrectID(sid)
            if rsid != sid:
                msg ='Subject: %s corrected to %s' % (sid,rsid)
                logging.warning(msg)
        else:
            rsid = sid
        return rsid

    def _loadData(self):
        """
        Detects if csv or Excel and loads self.data
        See http://www.python-excel.org/ for how pandas reads excel files
        :return:
        """
        try:
            if self.ftype =='.xlsx':
                #use openpyxl
                wb = load_workbook(filename=self.datafile)
                ws = wb[wb.sheetnames[self.sheet]]
                df = pandas.DataFrame(ws.values)

                if self.header is not None:
                    cols = df.iloc[self.header:self.header+1].values[0]
                    colnames = []
                    for i in range(len(cols)):
                        if cols[i] is None:
                            if df.iloc[i][0] is not None:
                                colnames.append(df.iloc[i][0])
                            else:
                                colnames.append(str(i))
                        else:
                            colnames.append(cols[i].replace(" ","").lower()+"_"+str(i))
                    df.columns = colnames
                if self.skiplines is not None and self.skiplines > 0:
                    df = df.drop(df.index[0:self.skiplines+1])
                self.data = df.reindex()

            elif self.ftype == '.xls':
                # older Excel version use xlrd
                if self.header is None:
                    self.data = pandas.read_excel(self.datafile, skiprows=self.skiplines, sheet_name=self.sheet, skip_blank_lines=True)
                else:
                    self.data = pandas.read_excel(self.datafile, skiprows=self.skiplines, sheet_name=self.sheet,
                                                  skip_blank_lines=True, header=self.header)
            elif self.ftype == '.csv':
                self.data = pandas.read_csv(self.datafile, skip_blank_lines=True)
            else:
                self.data = None
            if self.data is not None:
                msg = 'Data loaded from %s' % self.datafile
                logging.info(msg)
                print(msg)
                self.data.dropna(how="all", axis=0, inplace=True)  # cleanup if rows are all NaN
                self.data.fillna("")  # replace remaining NaNs with empty string

            else:
                raise ValueError('No data to load')
        except Exception as e:
            raise e

    def sortSubjects(self, subjectfield='ID'):
        '''
            Sort data into subjects by participant ID
             - this should be overwritten if the data is organized differently
        '''
        self.subjects = dict()
        if self.data is not None:
            if subjectfield not in self.data.columns:
                raise ValueError('Subject ID field not present: ', subjectfield)
            ids = self.data[subjectfield].unique()
            for sid in ids:
                if len(str(sid)) == 6:
                    sidkey = self.__checkSID(sid)
                    self.subjects[sidkey] = self.data[self.data[subjectfield] == sid]
            msg = 'Subjects loaded=%d' % len(self.subjects)
            print(msg)

    def formatDobNumber(self, orig):
        """
        Reformats DOB string from Excel data float to yyyy-mm-dd
        """
        dateoffset = 693594
        dt = datetime.fromordinal(dateoffset + int(orig))
        return dt.strftime("%Y-%m-%d")

    def formatCondensedDate(self, orig):
        """
        Reformats date number from Excel to yyyymmdd
        """
        dateoffset = 693594
        dt = datetime.fromordinal(dateoffset + int(orig))
        return dt.strftime("%Y%m%d")


    def getPrefix(self):
        prefix=None
        if self.info is not None:
            prefix = self.info['prefix']
        return prefix

    def getxsd(self):
        xsd = None
        if self.info is not None:
            xsd = self.info['xsitype']
        return xsd


def convertExcelDate(orig):
    """
    Reformats date number from Excel to datetime
    """
    # from datetime import datetime
    dateoffset = 693594
    dt = datetime.fromordinal(dateoffset + int(orig))
    return dt

def stripspaces(row,column):
    """
    Strips out whitespace before, within, after a value in column of row
    :param row:
    :param column:
    :return:
    """
    val = str(row[column])
    return val.replace(" ",'')